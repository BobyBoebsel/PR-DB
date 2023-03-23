import mysql.connector
from openpyxl import load_workbook

def readxlsx():
    wb=load_workbook(filename='PR-Nr.xlsx')
    sheet_ranges=wb['PR-Nr.']
    i=4
    mylist=[]
    while(i <= sheet_ranges.max_row):
        if sheet_ranges['F' + str((i+1))].value is not None:
            dict = {'PRNR':sheet_ranges['B' + str(i)].value,'FAM': sheet_ranges['C' + str(i)].value, 'TEXT': sheet_ranges['F' + str(i)].value + " " + sheet_ranges['F' + str((i+1))].value}
        else:
            dict = {'PRNR':sheet_ranges['B' + str(i)].value,'FAM': sheet_ranges['C' + str(i)].value, 'TEXT': sheet_ranges['F' + str(i)].value }
        mylist.append(dict)
        i+=2
    return mylist

def inssql(data):
    mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="1234",
    database="pr"
    )
    cursor = mydb.cursor()
    for dict in data:
        fields = str(list(dict.keys()))[1:-1]
        values = (str(list(dict.values()))[1:-1])
        sql = 'INSERT INTO pr.PR (' + fields.replace("'","") + ') VALUES (' + values + ')'
        cursor.execute(sql)
        mydb.commit()
        print(dict, "record inserted.")
    mydb.close()

if __name__ == "__main__":
    x=readxlsx()
    inssql(x)
